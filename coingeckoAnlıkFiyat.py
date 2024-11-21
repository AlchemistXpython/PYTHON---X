# pyinstaller --onefile --noconsole --exclude-module tkinter --exclude-module matplotlib --clean coingeckoAnlıkFiyat.py
# pyinstaller --onefile --noconsole --upx-dir /path/to/upx coingeckoAnlıkFiyat.py   

#v9  ekrana da bilgi ver  coin sıralamasının verilen listeye göre olması---------------------------------------------------------------------------------------------

import requests
import pandas as pd
from tkinter import Tk, Label, Entry, Button, Listbox, messagebox, Frame, Toplevel
from openpyxl import load_workbook
import logging

# Logging ayarları
logging.basicConfig(filename="crypto_app.log", level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# Coin listesi
coin_list = []
excel_file = "crypto_data.xlsx"

def add_coin():
    """Coin ismini listeye ekler."""
    coin_name = entry.get().strip().lower()
    if not coin_name:
        messagebox.showwarning("Uyarı", "Lütfen bir coin ismi girin.")
        return

    if coin_name in coin_list:
        messagebox.showwarning("Uyarı", "Bu coin zaten listede mevcut.")
        return

    coin_list.append(coin_name)
    coin_listbox.insert("end", coin_name)
    entry.delete(0, "end")
    logging.info(f"Coin eklendi: {coin_name}")

def fetch_crypto_data():
    """Listedeki coinler için veri çeker ve Excel'e yaz."""
    if not coin_list:
        messagebox.showwarning("Uyarı", "Önce coin ekleyin.")
        return

    api_url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {
        "vs_currency": "usd",
        "ids": ",".join(coin_list),  # Listedeki coinleri API'ye gönder
        "order": "market_cap_desc",
        "sparkline": False
    }

    try:
        response = requests.get(api_url, params=params)
        response.raise_for_status()  # HTTP hatalarını kontrol et

        data = response.json()
        if not data:
            messagebox.showwarning("Uyarı", "API'den geçerli bir veri alınamadı. Coin isimlerini kontrol edin.")
            logging.warning("API'den geçerli veri alınamadı.")
            return

        formatted_data = []
        current_date = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")  # Tarih ve saat

        for coin in data:
            formatted_data.append({
                "Tarih": current_date,
                "Coin": coin.get("name", "Bilinmiyor"),
                "Değişim (24h)": f'{coin.get("price_change_percentage_24h", 0):.2f}%',
                "Fiyat (USD)": f'${coin.get("current_price", 0):.6f}',
                "Hacim (24h)": f'${coin.get("total_volume", 0):,}',
                "Piyasa Değeri (USD)": f'${coin.get("market_cap", 0):,}',  # Piyasa değeri
                "Dolaşımdaki Arz": f'{coin.get("circulating_supply", 0):,}',
                "Toplam Arz": coin.get("total_supply", "Bilinmiyor")
            })

        df = pd.DataFrame(formatted_data)
        logging.info(f"İşlenen veri: {df}")

        # Excel'e veri yazma
        try:
            with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                sheet = writer.sheets.get("Sheet1", None)
                startrow = sheet.max_row if sheet else 0
                if startrow == 0:  # Başlıkları yalnızca bir kez yaz
                    df.to_excel(writer, sheet_name="Sheet1", index=False)
                else:
                    df.to_excel(writer, sheet_name="Sheet1", index=False, header=False, startrow=startrow)
                messagebox.showinfo("Başarılı", "Veriler başarıyla Excel dosyasına eklendi.")
                logging.info("Veriler Excel dosyasına başarıyla yazıldı.")
        except FileNotFoundError:
            with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Sheet1", index=False)
            messagebox.showinfo("Başarılı", "Yeni bir Excel dosyası oluşturuldu ve veriler yazıldı.")
            logging.info("Yeni Excel dosyası oluşturuldu ve veriler yazıldı.")

    except requests.exceptions.HTTPError as e:
        if response.status_code == 429:
            messagebox.showerror("Hata", "API limitine ulaşıldı. Lütfen daha sonra tekrar deneyin.")
        else:
            messagebox.showerror("Hata", f"API hatası: {response.status_code}")
        logging.error(f"API hatası: {e}")
    except Exception as e:
        messagebox.showerror("Hata", f"Veri alınırken bir hata oluştu: {e}")
        logging.error(f"Veri alınırken bir hata oluştu: {e}")

def search_and_display():
    """Verilen listeye göre coinleri sıralar ve mevcut pencerede tabloyu günceller."""
    if not coin_list:
        messagebox.showwarning("Uyarı", "Önce coin ekleyin.")
        return

    api_url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {
        "vs_currency": "usd",
        "ids": ",".join(coin_list),  # Listedeki coinleri API'ye gönder
        "order": "market_cap_desc",
        "sparkline": False
    }

    try:
        response = requests.get(api_url, params=params)
        response.raise_for_status()  # HTTP hatalarını kontrol et

        data = response.json()
        if not data:
            messagebox.showwarning("Uyarı", "API'den geçerli bir veri alınamadı.")
            logging.warning("API'den geçerli veri alınamadı.")
            return

        formatted_data = []
        for coin in data:
            formatted_data.append({
                "Coin": coin.get("name", "Bilinmiyor"),
                "Fiyat (USD)": f'${coin.get("current_price", 0):.6f}',
                "Değişim (24h)": f'{coin.get("price_change_percentage_24h", 0):.2f}%',
                "Hacim (24h)": f'${coin.get("total_volume", 0):,}'
            })

        # Pandas ile tablo oluşturma
        df = pd.DataFrame(formatted_data)
        logging.info(f"Arama sonucu: {df}")

        # Küçük harfe çevirip sıralama yapıyoruz
        df['Coin'] = df['Coin'].apply(lambda x: x.lower())  # Coin isimlerini küçük harfe çeviriyoruz
        df['Değişim (24h)'] = df['Değişim (24h)'].apply(lambda x: float(x.strip('%')))  # Değişim yüzdesini sayıya çeviriyoruz

        # Değişim yüzdesine göre sıralıyoruz
        df_sorted = df.sort_values(by='Değişim (24h)', ascending=False)

        # Eğer yeni pencere yoksa açalım
        if not hasattr(search_and_display, "new_window") or not search_and_display.new_window.winfo_exists():
            search_and_display.new_window = Toplevel(root)
            search_and_display.new_window.title("Coin Arama Sonuçları")
            search_and_display.new_window.geometry("600x400")

            # Yenile Butonu (üst solda olacak)
            def refresh_data():
                """Veriyi yenile ve sıralı göster."""
                search_and_display()  # Veriyi güncelle

            Button(search_and_display.new_window, text="Yenile", command=refresh_data).grid(row=0, column=0, padx=5, pady=5, sticky='w')

            # Başlıkları ekleyelim
            Label(search_and_display.new_window, text="Coin İsmi", width=20, anchor='w').grid(row=1, column=0, padx=5, pady=5)
            Label(search_and_display.new_window, text="Fiyat (USD)", width=20, anchor='c').grid(row=1, column=1, padx=5, pady=5)
            Label(search_and_display.new_window, text="Değişim (24h)", width=20, anchor='c').grid(row=1, column=2, padx=5, pady=5)
            Label(search_and_display.new_window, text="Hacim (24h)", width=20, anchor='c').grid(row=1, column=3, padx=5, pady=5)

        # Verileri tablo olarak ekleyelim
        for idx, row in df_sorted.iterrows():
            Label(search_and_display.new_window, text=row['Coin'], width=20, anchor='w').grid(row=idx+2, column=0, padx=5, pady=5)
            Label(search_and_display.new_window, text=row['Fiyat (USD)'], width=20, anchor='c').grid(row=idx+2, column=1, padx=5, pady=5)
            Label(search_and_display.new_window, text=f'{row["Değişim (24h)"]}%', width=20, anchor='c').grid(row=idx+2, column=2, padx=5, pady=5)
            Label(search_and_display.new_window, text=row['Hacim (24h)'], width=20, anchor='c').grid(row=idx+2, column=3, padx=5, pady=5)

        logging.info("Tablo yeni pencerede görüntülendi.")

    except requests.exceptions.HTTPError as e:
        messagebox.showerror("Hata", f"API hatası: {e}")
        logging.error(f"API hatası: {e}")
    except Exception as e:
        messagebox.showerror("Hata", f"Veri alınırken bir hata oluştu: {e}")
        logging.error(f"Veri alınırken bir hata oluştu: {e}")

def load_from_excel():
    """Excel'deki coin isimlerini listeye ekler."""
    try:
        workbook = load_workbook(excel_file)
        if "Coin" in workbook.sheetnames:
            sheet = workbook["Coin"]
            for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                coin_name = row[0]
                if coin_name and coin_name.strip().lower() not in coin_list:  # Boşlukları temizle
                    coin_list.append(coin_name.strip().lower())
                    coin_listbox.insert("end", coin_name.strip().lower())
            messagebox.showinfo("Başarılı", "Excel'den coin isimleri yüklendi.")
            logging.info("Excel'den coin isimleri başarıyla yüklendi.")
        else:
            messagebox.showwarning("Hata", "'Coin' isimli sayfa bulunamadı.")
            logging.warning("'Coin' isimli sayfa bulunamadı.")
    except FileNotFoundError:
        messagebox.showerror("Hata", f"{excel_file} dosyası bulunamadı.")
        logging.error(f"Excel dosyası bulunamadı: {excel_file}")
    except Exception as e:
        messagebox.showerror("Hata", f"Excel yüklenirken bir hata oluştu: {e}")
        logging.error(f"Excel yüklenirken bir hata oluştu: {e}")

def clear_list():
    """Coin listesini temizler."""
    coin_list.clear()
    coin_listbox.delete(0, "end")
    logging.info("Coin listesi temizlendi.")

# GUI oluşturma
root = Tk()
root.title("Kripto Veri Çekme")
root.geometry("400x650")

Label(root, text="Coin İsmini Gir ve Ekle:").pack(pady=5)
entry = Entry(root, width=40)
entry.pack(pady=5)

Button(root, text="Ekle", command=add_coin).pack(pady=5)

Label(root, text="Eklenen Coinler:").pack(pady=10)
coin_listbox = Listbox(root, width=50, height=15)
coin_listbox.pack(pady=5)

# Button'ları ortalamak için Frame kullanalım
button_frame = Frame(root)
button_frame.pack(pady=20)

Button(button_frame, text="Veri Çek ve Sheet1'e Yaz", command=fetch_crypto_data).pack(side="left", padx=5)
Button(button_frame, text="Yeni Sayfa ve Ara", command=search_and_display).pack(side="left", padx=5)

Button(root, text="Listeyi Temizle", command=clear_list).pack(pady=5)
Button(root, text="Excel'den Yükle", command=load_from_excel).pack(pady=10)

root.mainloop()







#v8  ekrana da bilgi ver  ve çıkan datanın hizalanması---------------------------------------------------------------------------------------------

# import requests
# import pandas as pd
# from tkinter import Tk, Label, Entry, Button, Listbox, messagebox, Frame, Toplevel
# from openpyxl import load_workbook
# import logging

# # Logging ayarları
# logging.basicConfig(filename="crypto_app.log", level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# # Coin listesi
# coin_list = []
# excel_file = "crypto_data.xlsx"

# def add_coin():
#     """Coin ismini listeye ekler."""
#     coin_name = entry.get().strip().lower()
#     if not coin_name:
#         messagebox.showwarning("Uyarı", "Lütfen bir coin ismi girin.")
#         return

#     if coin_name in coin_list:
#         messagebox.showwarning("Uyarı", "Bu coin zaten listede mevcut.")
#         return

#     coin_list.append(coin_name)
#     coin_listbox.insert("end", coin_name)
#     entry.delete(0, "end")
#     logging.info(f"Coin eklendi: {coin_name}")

# def fetch_crypto_data():
#     """Listedeki coinler için veri çeker ve Excel'e yaz."""
#     if not coin_list:
#         messagebox.showwarning("Uyarı", "Önce coin ekleyin.")
#         return

#     api_url = "https://api.coingecko.com/api/v3/coins/markets"
#     params = {
#         "vs_currency": "usd",
#         "ids": ",".join(coin_list),  # Listedeki coinleri API'ye gönder
#         "order": "market_cap_desc",
#         "sparkline": False
#     }

#     try:
#         response = requests.get(api_url, params=params)
#         response.raise_for_status()  # HTTP hatalarını kontrol et

#         data = response.json()
#         if not data:
#             messagebox.showwarning("Uyarı", "API'den geçerli bir veri alınamadı. Coin isimlerini kontrol edin.")
#             logging.warning("API'den geçerli veri alınamadı.")
#             return

#         formatted_data = []
#         current_date = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")  # Tarih ve saat

#         for coin in data:
#             formatted_data.append({
#                 "Tarih": current_date,
#                 "Coin": coin.get("name", "Bilinmiyor"),
#                 "Değişim (24h)": f'{coin.get("price_change_percentage_24h", 0):.2f}%',
#                 "Fiyat (USD)": f'${coin.get("current_price", 0):.6f}',
#                 "Hacim (24h)": f'${coin.get("total_volume", 0):,}',
#                 "Piyasa Değeri (USD)": f'${coin.get("market_cap", 0):,}',  # Piyasa değeri
#                 "Dolaşımdaki Arz": f'{coin.get("circulating_supply", 0):,}',
#                 "Toplam Arz": coin.get("total_supply", "Bilinmiyor")
#             })

#         df = pd.DataFrame(formatted_data)
#         logging.info(f"İşlenen veri: {df}")

#         # Excel'e veri yazma
#         try:
#             with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#                 sheet = writer.sheets.get("Sheet1", None)
#                 startrow = sheet.max_row if sheet else 0
#                 if startrow == 0:  # Başlıkları yalnızca bir kez yaz
#                     df.to_excel(writer, sheet_name="Sheet1", index=False)
#                 else:
#                     df.to_excel(writer, sheet_name="Sheet1", index=False, header=False, startrow=startrow)
#                 messagebox.showinfo("Başarılı", "Veriler başarıyla Excel dosyasına eklendi.")
#                 logging.info("Veriler Excel dosyasına başarıyla yazıldı.")
#         except FileNotFoundError:
#             with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
#                 df.to_excel(writer, sheet_name="Sheet1", index=False)
#             messagebox.showinfo("Başarılı", "Yeni bir Excel dosyası oluşturuldu ve veriler yazıldı.")
#             logging.info("Yeni Excel dosyası oluşturuldu ve veriler yazıldı.")

#     except requests.exceptions.HTTPError as e:
#         if response.status_code == 429:
#             messagebox.showerror("Hata", "API limitine ulaşıldı. Lütfen daha sonra tekrar deneyin.")
#         else:
#             messagebox.showerror("Hata", f"API hatası: {response.status_code}")
#         logging.error(f"API hatası: {e}")
#     except Exception as e:
#         messagebox.showerror("Hata", f"Veri alınırken bir hata oluştu: {e}")
#         logging.error(f"Veri alınırken bir hata oluştu: {e}")

# def search_and_display():
#     """Belirtilen coinleri arar ve yeni sayfada tablo olarak gösterir."""
#     if not coin_list:
#         messagebox.showwarning("Uyarı", "Önce coin ekleyin.")
#         return

#     api_url = "https://api.coingecko.com/api/v3/coins/markets"
#     params = {
#         "vs_currency": "usd",
#         "ids": ",".join(coin_list),  # Listedeki coinleri API'ye gönder
#         "order": "market_cap_desc",
#         "sparkline": False
#     }

#     try:
#         response = requests.get(api_url, params=params)
#         response.raise_for_status()  # HTTP hatalarını kontrol et

#         data = response.json()
#         if not data:
#             messagebox.showwarning("Uyarı", "API'den geçerli bir veri alınamadı.")
#             logging.warning("API'den geçerli veri alınamadı.")
#             return

#         formatted_data = []
#         for coin in data:
#             formatted_data.append({
#                 "Coin": coin.get("name", "Bilinmiyor"),
#                 "Fiyat (USD)": f'${coin.get("current_price", 0):.6f}',
#                 "Değişim (24h)": f'{coin.get("price_change_percentage_24h", 0):.2f}%',
#                 "Hacim (24h)": f'${coin.get("total_volume", 0):,}'
#             })

#         # Pandas ile tablo oluşturma
#         df = pd.DataFrame(formatted_data)
#         logging.info(f"Arama sonucu: {df}")

#         # Yeni pencere açma
#         new_window = Toplevel(root)
#         new_window.title("Coin Arama Sonuçları")
#         new_window.geometry("600x400")

#         # Başlıkları ekleyelim
#         Label(new_window, text="Coin İsmi", width=20, anchor='w').grid(row=0, column=0, padx=5, pady=5)
#         Label(new_window, text="Fiyat (USD)", width=20, anchor='c').grid(row=0, column=1, padx=5, pady=5)
#         Label(new_window, text="Değişim (24h)", width=20, anchor='c').grid(row=0, column=2, padx=5, pady=5)
#         Label(new_window, text="Hacim (24h)", width=20, anchor='c').grid(row=0, column=3, padx=5, pady=5)

#         # Verileri tablo olarak ekleyelim
#         for idx, row in df.iterrows():
#             Label(new_window, text=row['Coin'], width=20, anchor='w').grid(row=idx+1, column=0, padx=5, pady=5)
#             Label(new_window, text=row['Fiyat (USD)'], width=20, anchor='c').grid(row=idx+1, column=1, padx=5, pady=5)
#             Label(new_window, text=row['Değişim (24h)'], width=20, anchor='c').grid(row=idx+1, column=2, padx=5, pady=5)
#             Label(new_window, text=row['Hacim (24h)'], width=20, anchor='c').grid(row=idx+1, column=3, padx=5, pady=5)

#         logging.info("Tablo yeni pencerede görüntülendi.")

#     except requests.exceptions.HTTPError as e:
#         messagebox.showerror("Hata", f"API hatası: {e}")
#         logging.error(f"API hatası: {e}")
#     except Exception as e:
#         messagebox.showerror("Hata", f"Veri alınırken bir hata oluştu: {e}")
#         logging.error(f"Veri alınırken bir hata oluştu: {e}")

# def load_from_excel():
#     """Excel'deki coin isimlerini listeye ekler."""
#     try:
#         workbook = load_workbook(excel_file)
#         if "Coin" in workbook.sheetnames:
#             sheet = workbook["Coin"]
#             for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
#                 coin_name = row[0]
#                 if coin_name and coin_name.strip().lower() not in coin_list:  # Boşlukları temizle
#                     coin_list.append(coin_name.strip().lower())
#                     coin_listbox.insert("end", coin_name.strip().lower())
#             messagebox.showinfo("Başarılı", "Excel'den coin isimleri yüklendi.")
#             logging.info("Excel'den coin isimleri başarıyla yüklendi.")
#         else:
#             messagebox.showwarning("Hata", "'Coin' isimli sayfa bulunamadı.")
#             logging.warning("'Coin' isimli sayfa bulunamadı.")
#     except FileNotFoundError:
#         messagebox.showerror("Hata", f"{excel_file} dosyası bulunamadı.")
#         logging.error(f"Excel dosyası bulunamadı: {excel_file}")
#     except Exception as e:
#         messagebox.showerror("Hata", f"Excel yüklenirken bir hata oluştu: {e}")
#         logging.error(f"Excel yüklenirken bir hata oluştu: {e}")

# def clear_list():
#     """Coin listesini temizler."""
#     coin_list.clear()
#     coin_listbox.delete(0, "end")
#     logging.info("Coin listesi temizlendi.")

# # GUI oluşturma
# root = Tk()
# root.title("Kripto Veri Çekme")
# root.geometry("400x650")

# Label(root, text="Coin İsmini Gir ve Ekle:").pack(pady=5)
# entry = Entry(root, width=40)
# entry.pack(pady=5)

# Button(root, text="Ekle", command=add_coin).pack(pady=5)

# Label(root, text="Eklenen Coinler:").pack(pady=10)
# coin_listbox = Listbox(root, width=50, height=15)
# coin_listbox.pack(pady=5)

# # Button'ları ortalamak için Frame kullanalım
# button_frame = Frame(root)
# button_frame.pack(pady=20)

# Button(button_frame, text="Veri Çek ve Sheet1'e Yaz", command=fetch_crypto_data).pack(side="left", padx=5)
# Button(button_frame, text="Yeni Sayfa ve Ara", command=search_and_display).pack(side="left", padx=5)

# Button(root, text="Listeyi Temizle", command=clear_list).pack(pady=5)
# Button(root, text="Excel'den Yükle", command=load_from_excel).pack(pady=10)

# root.mainloop()



#v7  ekrana da bilgi ver ---------------------------------------------------------------------------------------------

# import requests
# import pandas as pd
# from tkinter import Tk, Label, Entry, Button, Listbox, messagebox, Frame  # Frame'i ekledik
# from openpyxl import load_workbook
# import logging

# # Logging ayarları
# logging.basicConfig(filename="crypto_app.log", level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# # Coin listesi
# coin_list = []
# excel_file = "crypto_data.xlsx"

# def add_coin():
#     """Coin ismini listeye ekler."""
#     coin_name = entry.get().strip().lower()
#     if not coin_name:
#         messagebox.showwarning("Uyarı", "Lütfen bir coin ismi girin.")
#         return

#     if coin_name in coin_list:
#         messagebox.showwarning("Uyarı", "Bu coin zaten listede mevcut.")
#         return

#     coin_list.append(coin_name)
#     coin_listbox.insert("end", coin_name)
#     entry.delete(0, "end")
#     logging.info(f"Coin eklendi: {coin_name}")

# def fetch_crypto_data():
#     """Listedeki coinler için veri çeker ve Excel'e yaz."""
#     if not coin_list:
#         messagebox.showwarning("Uyarı", "Önce coin ekleyin.")
#         return

#     api_url = "https://api.coingecko.com/api/v3/coins/markets"
#     params = {
#         "vs_currency": "usd",
#         "ids": ",".join(coin_list),  # Listedeki coinleri API'ye gönder
#         "order": "market_cap_desc",
#         "sparkline": False
#     }

#     try:
#         response = requests.get(api_url, params=params)
#         response.raise_for_status()  # HTTP hatalarını kontrol et

#         data = response.json()
#         if not data:
#             messagebox.showwarning("Uyarı", "API'den geçerli bir veri alınamadı. Coin isimlerini kontrol edin.")
#             logging.warning("API'den geçerli veri alınamadı.")
#             return

#         formatted_data = []
#         current_date = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")  # Tarih ve saat

#         for coin in data:
#             formatted_data.append({
#                 "Tarih": current_date,
#                 "Coin": coin.get("name", "Bilinmiyor"),
#                 "Değişim (24h)": f'{coin.get("price_change_percentage_24h", 0):.2f}%',
#                 "Fiyat (USD)": f'${coin.get("current_price", 0):.6f}',
#                 "Hacim (24h)": f'${coin.get("total_volume", 0):,}',
#                 "Piyasa Değeri (USD)": f'${coin.get("market_cap", 0):,}',  # Piyasa değeri
#                 "Dolaşımdaki Arz": f'{coin.get("circulating_supply", 0):,}',
#                 "Toplam Arz": coin.get("total_supply", "Bilinmiyor")
#             })

#         df = pd.DataFrame(formatted_data)
#         logging.info(f"İşlenen veri: {df}")

#         # Excel'e veri yazma
#         try:
#             with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#                 sheet = writer.sheets.get("Sheet1", None)
#                 startrow = sheet.max_row if sheet else 0
#                 if startrow == 0:  # Başlıkları yalnızca bir kez yaz
#                     df.to_excel(writer, sheet_name="Sheet1", index=False)
#                 else:
#                     df.to_excel(writer, sheet_name="Sheet1", index=False, header=False, startrow=startrow)
#                 messagebox.showinfo("Başarılı", "Veriler başarıyla Excel dosyasına eklendi.")
#                 logging.info("Veriler Excel dosyasına başarıyla yazıldı.")
#         except FileNotFoundError:
#             with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
#                 df.to_excel(writer, sheet_name="Sheet1", index=False)
#             messagebox.showinfo("Başarılı", "Yeni bir Excel dosyası oluşturuldu ve veriler yazıldı.")
#             logging.info("Yeni Excel dosyası oluşturuldu ve veriler yazıldı.")

#     except requests.exceptions.HTTPError as e:
#         if response.status_code == 429:
#             messagebox.showerror("Hata", "API limitine ulaşıldı. Lütfen daha sonra tekrar deneyin.")
#         else:
#             messagebox.showerror("Hata", f"API hatası: {response.status_code}")
#         logging.error(f"API hatası: {e}")
#     except Exception as e:
#         messagebox.showerror("Hata", f"Veri alınırken bir hata oluştu: {e}")
#         logging.error(f"Veri alınırken bir hata oluştu: {e}")

# def search_and_display():
#     """Belirtilen coinleri arar ve ekranda tablo olarak gösterir."""
#     if not coin_list:
#         messagebox.showwarning("Uyarı", "Önce coin ekleyin.")
#         return

#     api_url = "https://api.coingecko.com/api/v3/coins/markets"
#     params = {
#         "vs_currency": "usd",
#         "ids": ",".join(coin_list),  # Listedeki coinleri API'ye gönder
#         "order": "market_cap_desc",
#         "sparkline": False
#     }

#     try:
#         response = requests.get(api_url, params=params)
#         response.raise_for_status()  # HTTP hatalarını kontrol et

#         data = response.json()
#         if not data:
#             messagebox.showwarning("Uyarı", "API'den geçerli bir veri alınamadı.")
#             logging.warning("API'den geçerli veri alınamadı.")
#             return

#         formatted_data = []
#         for coin in data:
#             formatted_data.append({
#                 "Coin": coin.get("name", "Bilinmiyor"),
#                 "Fiyat (USD)": f'${coin.get("current_price", 0):.6f}',
#                 "Değişim (24h)": f'{coin.get("price_change_percentage_24h", 0):.2f}%',
#                 "Hacim (24h)": f'${coin.get("total_volume", 0):,}'
#             })

#         # Pandas ile tablo oluşturma
#         df = pd.DataFrame(formatted_data)
#         logging.info(f"Arama sonucu: {df}")

#         # Tabloyu ekranda göstermek
#         display_text = df.to_string(index=False)
#         messagebox.showinfo("Arama Sonucu", display_text)
#         logging.info("Tablo ekranda görüntülendi.")

#     except requests.exceptions.HTTPError as e:
#         messagebox.showerror("Hata", f"API hatası: {e}")
#         logging.error(f"API hatası: {e}")
#     except Exception as e:
#         messagebox.showerror("Hata", f"Veri alınırken bir hata oluştu: {e}")
#         logging.error(f"Veri alınırken bir hata oluştu: {e}")

# def load_from_excel():
#     """Excel'deki coin isimlerini listeye ekler."""
#     try:
#         workbook = load_workbook(excel_file)
#         if "Coin" in workbook.sheetnames:
#             sheet = workbook["Coin"]
#             for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
#                 coin_name = row[0]
#                 if coin_name and coin_name.strip().lower() not in coin_list:  # Boşlukları temizle
#                     coin_list.append(coin_name.strip().lower())
#                     coin_listbox.insert("end", coin_name.strip().lower())
#             messagebox.showinfo("Başarılı", "Excel'den coin isimleri yüklendi.")
#             logging.info("Excel'den coin isimleri başarıyla yüklendi.")
#         else:
#             messagebox.showwarning("Hata", "'Coin' isimli sayfa bulunamadı.")
#             logging.warning("'Coin' isimli sayfa bulunamadı.")
#     except FileNotFoundError:
#         messagebox.showerror("Hata", f"{excel_file} dosyası bulunamadı.")
#         logging.error(f"Excel dosyası bulunamadı: {excel_file}")
#     except Exception as e:
#         messagebox.showerror("Hata", f"Excel yüklenirken bir hata oluştu: {e}")
#         logging.error(f"Excel yüklenirken bir hata oluştu: {e}")

# def clear_list():
#     """Coin listesini temizler."""
#     coin_list.clear()
#     coin_listbox.delete(0, "end")
#     logging.info("Coin listesi temizlendi.")

# # GUI oluşturma
# root = Tk()
# root.title("Kripto Veri Çekme")
# root.geometry("400x650")

# Label(root, text="Coin İsmini Gir ve Ekle:").pack(pady=5)
# entry = Entry(root, width=40)
# entry.pack(pady=5)

# Button(root, text="Ekle", command=add_coin).pack(pady=5)

# Label(root, text="Eklenen Coinler:").pack(pady=10)
# coin_listbox = Listbox(root, width=50, height=15)
# coin_listbox.pack(pady=5)

# # Button'ları ortalamak için Frame kullanalım
# button_frame = Frame(root)
# button_frame.pack(pady=20)

# Button(button_frame, text="Veri Çek ve Sheet1'e Yaz", command=fetch_crypto_data).pack(side="left", padx=5)
# Button(button_frame, text="Yeni Sayfa ve Ara", command=search_and_display).pack(side="left", padx=5)

# Button(root, text="Listeyi Temizle", command=clear_list).pack(pady=5)
# Button(root, text="Excel'den Yükle", command=load_from_excel).pack(pady=10)

# root.mainloop()




#v6  Coin sayfasını olduğu gibi bırak---------------------------------------------------------------------------------------------

# import requests
# import pandas as pd
# from tkinter import Tk, Label, Entry, Button, Listbox, messagebox
# from openpyxl import load_workbook
# import logging

# # Logging ayarları
# logging.basicConfig(filename="crypto_app.log", level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# # Coin listesi
# coin_list = []
# excel_file = "crypto_data.xlsx"

# def add_coin():
#     """Coin ismini listeye ekler."""
#     coin_name = entry.get().strip().lower()
#     if not coin_name:
#         messagebox.showwarning("Uyarı", "Lütfen bir coin ismi girin.")
#         return

#     if coin_name in coin_list:
#         messagebox.showwarning("Uyarı", "Bu coin zaten listede mevcut.")
#         return

#     coin_list.append(coin_name)
#     coin_listbox.insert("end", coin_name)
#     entry.delete(0, "end")
#     logging.info(f"Coin eklendi: {coin_name}")

# def fetch_crypto_data():
#     """Listedeki coinler için veri çeker ve Excel'e yaz."""
#     if not coin_list:
#         messagebox.showwarning("Uyarı", "Önce coin ekleyin.")
#         return

#     api_url = "https://api.coingecko.com/api/v3/coins/markets"
#     params = {
#         "vs_currency": "usd",
#         "ids": ",".join(coin_list),  # Listedeki coinleri API'ye gönder
#         "order": "market_cap_desc",
#         "sparkline": False
#     }

#     try:
#         response = requests.get(api_url, params=params)
#         response.raise_for_status()  # HTTP hatalarını kontrol et

#         data = response.json()
#         if not data:
#             messagebox.showwarning("Uyarı", "API'den geçerli bir veri alınamadı. Coin isimlerini kontrol edin.")
#             logging.warning("API'den geçerli veri alınamadı.")
#             return

#         formatted_data = []
#         current_date = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")  # Tarih ve saat

#         for coin in data:
#             formatted_data.append({
#                 "Tarih": current_date,
#                 "Coin": coin.get("name", "Bilinmiyor"),
#                 "Değişim (24h)": f'{coin.get("price_change_percentage_24h", 0):.2f}%',
#                 "Fiyat (USD)": f'${coin.get("current_price", 0):.6f}',
#                 "Hacim (24h)": f'${coin.get("total_volume", 0):,}',
#                 "Piyasa Değeri (USD)": f'${coin.get("market_cap", 0):,}',  # Piyasa değeri
#                 "Dolaşımdaki Arz": f'{coin.get("circulating_supply", 0):,}',
#                 "Toplam Arz": coin.get("total_supply", "Bilinmiyor")
#             })

#         df = pd.DataFrame(formatted_data)
#         logging.info(f"İşlenen veri: {df}")

#         # Excel'e veri yazma
#         try:
#             with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#                 sheet = writer.sheets.get("Sheet1", None)
#                 startrow = sheet.max_row if sheet else 0
#                 if startrow == 0:  # Başlıkları yalnızca bir kez yaz
#                     df.to_excel(writer, sheet_name="Sheet1", index=False)
#                 else:
#                     df.to_excel(writer, sheet_name="Sheet1", index=False, header=False, startrow=startrow)
#                 messagebox.showinfo("Başarılı", "Veriler başarıyla Excel dosyasına eklendi.")
#                 logging.info("Veriler Excel dosyasına başarıyla yazıldı.")
#         except FileNotFoundError:
#             with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
#                 df.to_excel(writer, sheet_name="Sheet1", index=False)
#             messagebox.showinfo("Başarılı", "Yeni bir Excel dosyası oluşturuldu ve veriler yazıldı.")
#             logging.info("Yeni Excel dosyası oluşturuldu ve veriler yazıldı.")

#     except requests.exceptions.HTTPError as e:
#         if response.status_code == 429:
#             messagebox.showerror("Hata", "API limitine ulaşıldı. Lütfen daha sonra tekrar deneyin.")
#         else:
#             messagebox.showerror("Hata", f"API hatası: {response.status_code}")
#         logging.error(f"API hatası: {e}")
#     except Exception as e:
#         messagebox.showerror("Hata", f"Veri alınırken bir hata oluştu: {e}")
#         logging.error(f"Veri alınırken bir hata oluştu: {e}")

# def load_from_excel():
#     """Excel'deki coin isimlerini listeye ekler."""
#     try:
#         workbook = load_workbook(excel_file)
#         if "Coin" in workbook.sheetnames:
#             sheet = workbook["Coin"]
#             for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
#                 coin_name = row[0]
#                 if coin_name and coin_name.strip().lower() not in coin_list:  # Boşlukları temizle
#                     coin_list.append(coin_name.strip().lower())
#                     coin_listbox.insert("end", coin_name.strip().lower())
#             messagebox.showinfo("Başarılı", "Excel'den coin isimleri yüklendi.")
#             logging.info("Excel'den coin isimleri başarıyla yüklendi.")
#         else:
#             messagebox.showwarning("Hata", "'Coin' isimli sayfa bulunamadı.")
#             logging.warning("'Coin' isimli sayfa bulunamadı.")
#     except FileNotFoundError:
#         messagebox.showerror("Hata", f"{excel_file} dosyası bulunamadı.")
#         logging.error(f"Excel dosyası bulunamadı: {excel_file}")
#     except Exception as e:
#         messagebox.showerror("Hata", f"Excel yüklenirken bir hata oluştu: {e}")
#         logging.error(f"Excel yüklenirken bir hata oluştu: {e}")

# def clear_list():
#     """Coin listesini temizler."""
#     coin_list.clear()
#     coin_listbox.delete(0, "end")
#     logging.info("Coin listesi temizlendi.")

# # GUI oluşturma
# root = Tk()
# root.title("Kripto Veri Çekme")
# root.geometry("400x550")

# Label(root, text="Coin İsmini Gir ve Ekle:").pack(pady=5)
# entry = Entry(root, width=40)
# entry.pack(pady=5)

# Button(root, text="Ekle", command=add_coin).pack(pady=5)

# Label(root, text="Eklenen Coinler:").pack(pady=10)
# coin_listbox = Listbox(root, width=50, height=15)
# coin_listbox.pack(pady=5)

# Button(root, text="Veri Çek ve Sheet1'e Yaz", command=fetch_crypto_data).pack(pady=10)
# Button(root, text="Listeyi Temizle", command=clear_list).pack(pady=5)
# Button(root, text="Excel'den Yükle", command=load_from_excel).pack(pady=10)

# root.mainloop()





#v6.01---------------------------------------------------------------------------------------------------------------------------------------------
#coinmarketcap ---------------------------------------------------------------------------------------------------------------------------------------
# import requests
# import pandas as pd

# # CoinMarketCap API bilgileri
# api_key = "YOUR_API_KEY"  # CoinMarketCap API anahtarınızı buraya ekleyin
# api_url = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest"

# headers = {
#     "Accepts": "application/json",
#     "X-CMC_PRO_API_KEY": api_key,
# }

# params = {
#     "start": "1",
#     "limit": "10",
#     "convert": "USD"
# }

# def fetch_coinmarketcap_data():
#     try:
#         response = requests.get(api_url, headers=headers, params=params)
#         response.raise_for_status()
#         data = response.json()

#         # Verileri işleme
#         coins = data["data"]
#         formatted_data = []
#         for coin in coins:
#             formatted_data.append({
#                 "Coin": coin["name"],
#                 "Sembol": coin["symbol"],
#                 "Fiyat (USD)": f'{coin["quote"]["USD"]["price"]:.2f}',
#                 "Değişim (24h)": f'{coin["quote"]["USD"]["percent_change_24h"]:.2f}%',
#                 "Piyasa Değeri": f'{coin["quote"]["USD"]["market_cap"]:.2f}',
#                 "Hacim (24h)": f'{coin["quote"]["USD"]["volume_24h"]:.2f}'
#             })

#         df = pd.DataFrame(formatted_data)
#         print(df)
#     except Exception as e:
#         print(f"Hata: {e}")

# fetch_coinmarketcap_data()






#v6.02---------------------------------------------------------------------------------------------------------------------------------------------
#tradinview ------------------------------------------------------------------------------------------------------------------------------------
# from tvDatafeed import TvDatafeed, Interval

# # Kullanıcı bilgileri
# username = "your_username"  # TradingView kullanıcı adı
# password = "your_password"  # TradingView şifresi

# tv = TvDatafeed(username, password)

# # Veri çekme (örnek BTC/USDT)
# def fetch_tradingview_data():
#     try:
#         data = tv.get_hist(
#             symbol="BTCUSDT",
#             exchange="BINANCE",
#             interval=Interval.in_daily,
#             n_bars=100
#         )
#         print(data)
#     except Exception as e:
#         print(f"Hata: {e}")

# fetch_tradingview_data()




#v5 Coin sayfasını siliyordu----------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------
# import requests
# import pandas as pd
# from tkinter import Tk, Label, Entry, Button, Listbox, messagebox
# from openpyxl import load_workbook

# # Coin listesi
# coin_list = []
# excel_file = "crypto_data.xlsx"

# def add_coin():
#     """Coin ismini listeye ekler."""
#     coin_name = entry.get().strip().lower()
#     if coin_name and coin_name not in coin_list:
#         coin_list.append(coin_name)
#         coin_listbox.insert("end", coin_name)
#         entry.delete(0, "end")
#     else:
#         messagebox.showwarning("Uyarı", "Geçerli bir coin ismi girin veya zaten listede mevcut.")

# def fetch_crypto_data():
#     """Listedeki coinler için veri çeker ve Excel'e yazar."""
#     if not coin_list:
#         messagebox.showwarning("Uyarı", "Önce coin ekleyin.")
#         return

#     api_url = "https://api.coingecko.com/api/v3/coins/markets"
#     params = {
#         "vs_currency": "usd",
#         "ids": ",".join(coin_list),  # Listedeki coinleri API'ye gönder
#         "order": "market_cap_desc",
#         "sparkline": False
#     }

#     try:
#         response = requests.get(api_url, params=params)
#         response.raise_for_status()
#         data = response.json()

#         # Veriyi formatla
#         formatted_data = []
#         current_date = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")  # Tarih ve saat

#         for coin in data:
#             formatted_data.append({
#                 "Tarih": current_date,
#                 "Coin": coin["name"],
#                 "Değişim (24h)": f'{coin["price_change_percentage_24h"]:.2f}%',
#                 "Fiyat (USD)": f'${coin["current_price"]:.6f}',
#                 "Hacim (24h)": f'${coin["total_volume"]:,}',
#                 "Dolaşımdaki Arz": f'{coin["circulating_supply"]:,}',
#                 "Toplam Arz": coin.get("total_supply", "Bilinmiyor")  # Toplam arz boş olabilir
#             })

#         # Veriyi Excel'e yaz
#         df = pd.DataFrame(formatted_data)

#         # Dosya zaten varsa üzerine ekle
#         try:
#             existing_data = pd.read_excel(excel_file)
#             df = pd.concat([existing_data, df], ignore_index=True)
#         except FileNotFoundError:
#             pass  # Dosya yoksa sıfırdan yaz

#         df.to_excel(excel_file, index=False)
#         messagebox.showinfo("Başarılı", f"Veriler '{excel_file}' dosyasına yazıldı.")

#     except Exception as e:
#         messagebox.showerror("Hata", f"Veri alınırken bir hata oluştu: {e}")

# def load_from_excel():
#     """Excel'deki coin isimlerini listeye ekler."""
#     try:
#         # Excel dosyasını oku
#         workbook = load_workbook(excel_file)
#         if "Coin" in workbook.sheetnames:
#             sheet = workbook["Coin"]  # "Coin" isimli sayfayı seç
#             for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):  # A sütunundan değerleri al
#                 coin_name = row[0]
#                 if coin_name and coin_name.lower() not in coin_list:
#                     coin_list.append(coin_name.lower())
#                     coin_listbox.insert("end", coin_name.lower())
#             messagebox.showinfo("Başarılı", "Excel'den coin isimleri yüklendi.")
#         else:
#             messagebox.showwarning("Hata", "Excel dosyasında 'Coin' isimli sayfa bulunamadı.")
#     except FileNotFoundError:
#         messagebox.showerror("Hata", f"'{excel_file}' dosyası bulunamadı.")
#     except Exception as e:
#         messagebox.showerror("Hata", f"Excel yüklenirken bir hata oluştu: {e}")

# def clear_list():
#     """Coin listesini temizler."""
#     coin_list.clear()
#     coin_listbox.delete(0, "end")

# # GUI oluşturma
# root = Tk()
# root.title("Kripto Veri Çekme")
# root.geometry("500x500")

# Label(root, text="Coin İsmini Gir ve Ekle:").pack(pady=5)
# entry = Entry(root, width=40)
# entry.pack(pady=5)

# Button(root, text="Ekle", command=add_coin).pack(pady=5)

# Label(root, text="Eklenen Coinler:").pack(pady=10)
# coin_listbox = Listbox(root, width=50, height=15)
# coin_listbox.pack(pady=5)

# Button(root, text="Veri Çek ve Excel'e Yaz", command=fetch_crypto_data).pack(pady=10)
# Button(root, text="Listeyi Temizle", command=clear_list).pack(pady=5)
# Button(root, text="Excel'den Yükle", command=load_from_excel).pack(pady=10)

# root.mainloop()



#v4 listeye coin ismi ekleme-----------------------------------------------------------------------------------------------------------------

# import requests
# import pandas as pd
# from tkinter import Tk, Label, Entry, Button, Listbox, messagebox

# # Coin listesi
# coin_list = []
# excel_file = "crypto_data.xlsx"

# def add_coin():
#     """Coin ismini listeye ekler."""
#     coin_name = entry.get().strip().lower()
#     if coin_name and coin_name not in coin_list:
#         coin_list.append(coin_name)
#         coin_listbox.insert("end", coin_name)
#         entry.delete(0, "end")
#     else:
#         messagebox.showwarning("Uyarı", "Geçerli bir coin ismi girin veya zaten listede mevcut.")

# def fetch_crypto_data():
#     """Listedeki coinler için veri çeker ve Excel'e yazar."""
#     if not coin_list:
#         messagebox.showwarning("Uyarı", "Önce coin ekleyin.")
#         return

#     api_url = "https://api.coingecko.com/api/v3/coins/markets"
#     params = {
#         "vs_currency": "usd",
#         "ids": ",".join(coin_list),  # Listedeki coinleri API'ye gönder
#         "order": "market_cap_desc",
#         "sparkline": False
#     }

#     try:
#         response = requests.get(api_url, params=params)
#         response.raise_for_status()
#         data = response.json()

#         # Veriyi formatla
#         formatted_data = []
#         current_date = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")  # Tarih ve saat

#         for coin in data:
#             formatted_data.append({
#                 "Tarih": current_date,
#                 "Coin": coin["name"],
#                 "Değişim (24h)": f'{coin["price_change_percentage_24h"]:.2f}%',
#                 "Fiyat (USD)": f'${coin["current_price"]:.6f}',
#                 "Hacim (24h)": f'${coin["total_volume"]:,}',
#                 "Dolaşımdaki Arz": f'{coin["circulating_supply"]:,}',
#                 "Toplam Arz": coin.get("total_supply", "Bilinmiyor")  # Toplam arz boş olabilir
#             })

#         # Veriyi Excel'e yaz
#         df = pd.DataFrame(formatted_data)

#         # Dosya zaten varsa üzerine ekle
#         try:
#             existing_data = pd.read_excel(excel_file)
#             df = pd.concat([existing_data, df], ignore_index=True)
#         except FileNotFoundError:
#             pass  # Dosya yoksa sıfırdan yaz

#         df.to_excel(excel_file, index=False)
#         messagebox.showinfo("Başarılı", f"Veriler '{excel_file}' dosyasına yazıldı.")

#     except Exception as e:
#         messagebox.showerror("Hata", f"Veri alınırken bir hata oluştu: {e}")

# def clear_list():
#     """Coin listesini temizler."""
#     coin_list.clear()
#     coin_listbox.delete(0, "end")

# # GUI oluşturma
# root = Tk()
# root.title("Kripto Veri Çekme")
# root.geometry("500x400")

# Label(root, text="Coin İsmini Gir ve Ekle:").pack(pady=5)
# entry = Entry(root, width=40)
# entry.pack(pady=5)

# Button(root, text="Ekle", command=add_coin).pack(pady=5)

# Label(root, text="Eklenen Coinler:").pack(pady=10)
# coin_listbox = Listbox(root, width=50, height=10)
# coin_listbox.pack(pady=5)

# Button(root, text="Veri Çek ve Excel'e Yaz", command=fetch_crypto_data).pack(pady=10)
# Button(root, text="Listeyi Temizle", command=clear_list).pack(pady=5)

# root.mainloop()

#------------------------------------------------------------------------------------------------------------------------------------------------
#v2 tek coin ekleniyor

# import requests
# import pandas as pd
# from tkinter import Tk, Label, Entry, Button, messagebox

# def fetch_crypto_data(crypto_names):
#     api_url = "https://api.coingecko.com/api/v3/coins/markets"
#     params = {
#         "vs_currency": "usd",
#         "ids": ",".join(crypto_names),  # Girilen coin isimlerini API'ye gönder
#         "order": "market_cap_desc",
#         "sparkline": False
#     }
#     try:
#         response = requests.get(api_url, params=params)
#         response.raise_for_status()
#         data = response.json()
        
#         # Excel için formatlama
#         formatted_data = []
#         current_date = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")  # Tarih ve saat
        
#         for coin in data:
#             formatted_data.append({
#                 "Tarih": current_date,
#                 "Coin": coin["name"],
#                 "Değişim (24h)": f'{coin["price_change_percentage_24h"]:.2f}%',
#                 "Fiyat (USD)": f'${coin["current_price"]:.6f}',
#                 "Hacim (24h)": f'${coin["total_volume"]:,}',
#                 "Dolaşımdaki Arz": f'{coin["circulating_supply"]:,}',
#                 "Toplam Arz": coin.get("total_supply", "Bilinmiyor")  # Toplam arz boş olabilir
#             })
        
#         # Veriyi Excel'e yaz
#         df = pd.DataFrame(formatted_data)
#         output_file = "crypto_data.xlsx"
#         df.to_excel(output_file, index=False)
        
#         messagebox.showinfo("Başarılı", f"Veriler '{output_file}' dosyasına yazıldı.")
#     except Exception as e:
#         messagebox.showerror("Hata", f"Veri alınırken bir hata oluştu: {e}")

# def on_submit():
#     crypto_names = entry.get().strip().lower().split(",")  # Virgül ile ayrılan coin isimlerini al
#     fetch_crypto_data(crypto_names)

# # GUI oluşturma
# root = Tk()
# root.title("Kripto Veri Çekme")
# root.geometry("400x200")

# Label(root, text="Coin İsimlerini Gir (Virgülle Ayırın):").pack(pady=10)
# entry = Entry(root, width=50)
# entry.pack(pady=5)

# Button(root, text="Veri Çek ve Excel'e Yaz", command=on_submit).pack(pady=20)

# root.mainloop()



#v1--------------------------------------------------------------------------------------------------------------
# import requests

# # CoinGecko API URL'si
# api_url = "https://api.coingecko.com/api/v3/coins/markets"

# # Kripto paraların kimlikleri (CoinGecko'daki ID'ler)
# symbols = ["floki", "pepe", "arkham"]

# # API parametreleri
# params = {
#     "vs_currency": "usd",  # USD bazında fiyat
#     "ids": ",".join(symbols),  # FLOKI, PEPE ve ARKM bilgileri
#     "order": "market_cap_desc",
#     "per_page": len(symbols),
#     "page": 1,
#     "sparkline": False
# }

# def get_cryptocurrency_data():
#     try:
#         # API'den veri çekme
#         response = requests.get(api_url, params=params)
#         response.raise_for_status()  # HTTP hatalarını kontrol et
#         data = response.json()

#         # Her bir coin için bilgileri yazdırma
#         for coin in data:
#             print(f"{coin['name']} ({coin['symbol'].upper()}) Güncel Bilgiler:")
#             print(f"- Güncel Fiyat (USD): ${coin['current_price']}")
#             print(f"- Piyasa Değeri (USD): ${coin['market_cap']}")
#             print(f"- 24 Saatlik İşlem Hacmi (USD): ${coin['total_volume']}")
#             print(f"- 24 Saatlik Değişim (%): {coin['price_change_percentage_24h']:.2f}%")
#             print(f"- Dolaşımdaki Arz: {coin['circulating_supply']}")
#             print("-" * 50)

#     except Exception as e:
#         print(f"Hata: {e}")

# # Fonksiyonu çalıştır
# get_cryptocurrency_data()
